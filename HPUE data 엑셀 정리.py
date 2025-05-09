import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# 엑셀 파일 경로
file_path = r"C:\python\save\Keyonet\Ericsson_UL_Data.xlsx"
output_path = r"C:\python\save\Keyonet\Ericsson_UL_Data_Result.xlsx"

# 기존 엑셀 파일 열기
xls = pd.ExcelFile(file_path)
sheet_names = sorted(xls.sheet_names)

# 새 워크북 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

for sheet in sheet_names:
    print(f"▶️ 처리 중: {sheet}")
    try:
        df = pd.read_excel(file_path, sheet_name=sheet)

        # 필요한 열만 추출
        rsrp_df = df[["Local Time", "SSB RSRP Avg"]].dropna()
        power_df = df[["Local Time.1", "Power"]].dropna()
        tput_df = df[["Local Time.2", "NR PHY UL Tput (new+retx) [Mbps]"]].dropna()

        # 컬럼 이름 통일
        rsrp_df.columns = ["Time", "RSRP"]
        power_df.columns = ["Time", "Power"]
        tput_df.columns = ["Time", "Tput"]

        # 문자열을 시간으로 파싱 (에러는 무시)
        # 시간 문자열로 변환해서 날짜 제거 (엑셀에 시간만 보이도록)
        rsrp_df["Time"] = (
            pd.to_datetime(rsrp_df["Time"], format="%H:%M:%S:%f", errors="coerce")
            .dt.strftime("%H:%M:%S.%f")
            .str[:-3]
        )
        power_df["Time"] = (
            pd.to_datetime(power_df["Time"], format="%H:%M:%S:%f", errors="coerce")
            .dt.strftime("%H:%M:%S.%f")
            .str[:-3]
        )
        tput_df["Time"] = (
            pd.to_datetime(tput_df["Time"], format="%H:%M:%S:%f", errors="coerce")
            .dt.strftime("%H:%M:%S.%f")
            .str[:-3]
        )

        # 병합 후 정렬
        merged = power_df.merge(rsrp_df, on="Time", how="left").merge(
            tput_df, on="Time", how="left"
        )
        merged.sort_values(by="Time", inplace=True)

        # 시트 생성
        ws = wb.create_sheet(title=sheet)

        # 데이터프레임을 시트에 쓰기
        for row in dataframe_to_rows(merged, index=False, header=True):
            ws.append(row)

        # 차트 생성
        chart = LineChart()
        chart.title = f"{sheet} - RSRP / Power / Tput"
        chart.y_axis.title = "값"
        chart.x_axis.title = "시간"

    except Exception as e:
        print(f"❌ 시트 '{sheet}' 처리 중 오류: {e}")

# 새 파일로 저장
wb.save(output_path)
print(f"\n✅ 완료: {output_path}")
